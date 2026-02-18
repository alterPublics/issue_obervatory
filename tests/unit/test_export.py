"""Unit tests for the ContentExporter module.

Tests cover:
- CSV export: correct headers, correct row count, Danish character encoding (UTF-8, æøå)
- XLSX export: file produced, correct sheet name, Danish characters preserved
- JSON (NDJSON) export: valid JSON, all records present, raw_metadata intact
- GEXF actor export: node/edge structure, edge weight = shared term count (DQ-02 regression)
- GEXF actor export: edges grouped by shared search term, NOT by collection_run_id
- GEXF term export: term nodes, co-occurrence edges
- GEXF bipartite export: actor and term node types, term: prefix on IDs
- GEXF export: no duplicate nodes, no self-edges
- Empty dataset: each format returns empty-but-valid output
- content_hash deduplication: duplicate records not double-counted in networks

These tests run without a live database or network connection.
All DB calls are mocked at the session level; file output uses io.BytesIO.
"""

from __future__ import annotations

import io
import json
import os
import xml.etree.ElementTree as ET
from typing import Any

import pytest

# ---------------------------------------------------------------------------
# Env bootstrap
# ---------------------------------------------------------------------------

os.environ.setdefault("PSEUDONYMIZATION_SALT", "test-pseudonymization-salt-for-unit-tests")
os.environ.setdefault("SECRET_KEY", "test-secret-key-for-tests-only")
os.environ.setdefault("CREDENTIAL_ENCRYPTION_KEY", "dGVzdC1mZXJuZXQta2V5LTMyLWJ5dGVzLXBhZGRlZA==")

from issue_observatory.analysis.export import ContentExporter, _COLUMN_HEADERS, _FLAT_COLUMNS, _safe_str  # noqa: E402


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _make_record(
    platform: str = "bluesky",
    arena: str = "social_media",
    author_id: str = "author-abc",
    display_name: str = "Test Author",
    text: str = "Hello world",
    terms: list[str] | None = None,
    run_id: str = "run-001",
    raw_metadata: dict | None = None,
    likes: int | None = None,
    shares: int | None = None,
    comments: int | None = None,
    views: int | None = None,
) -> dict[str, Any]:
    """Build a minimal content record dict with sensible defaults."""
    return {
        "platform": platform,
        "arena": arena,
        "content_type": "post",
        "title": None,
        "text_content": text,
        "url": f"https://example.com/{author_id}",
        "author_display_name": display_name,
        "published_at": "2026-01-15T12:00:00+00:00",
        "views_count": views,
        "likes_count": likes,
        "shares_count": shares,
        "comments_count": comments,
        "language": "da",
        "collection_tier": "free",
        "search_terms_matched": terms or ["klimaforandringer"],
        "pseudonymized_author_id": author_id,
        "collection_run_id": run_id,
        "raw_metadata": raw_metadata or {"source": "test"},
    }


DANISH_TEXT = "Grøn omstilling er vigtig for velfærdsstaten og Ålborg"
DANISH_AUTHOR = "Søren Ærlighed-Øberg"

EXPORTER = ContentExporter()


# ---------------------------------------------------------------------------
# _safe_str helper
# ---------------------------------------------------------------------------


class TestSafeStr:
    def test_safe_str_returns_empty_string_for_none(self) -> None:
        """_safe_str(None) returns an empty string."""
        assert _safe_str(None) == ""

    def test_safe_str_joins_list_with_pipe(self) -> None:
        """_safe_str(['a', 'b']) returns 'a | b'."""
        assert _safe_str(["term1", "term2"]) == "term1 | term2"

    def test_safe_str_passes_through_string(self) -> None:
        """_safe_str('hello') returns 'hello' unchanged."""
        assert _safe_str("hello") == "hello"

    def test_safe_str_converts_int_to_string(self) -> None:
        """_safe_str(42) returns '42'."""
        assert _safe_str(42) == "42"

    def test_safe_str_preserves_danish_characters(self) -> None:
        """_safe_str preserves æ, ø, å without corruption."""
        result = _safe_str(DANISH_TEXT)
        assert "Grøn" in result
        assert "velfærdsstaten" in result
        assert "Ålborg" in result


# ---------------------------------------------------------------------------
# CSV export
# ---------------------------------------------------------------------------


class TestExportCsv:
    @pytest.mark.asyncio
    async def test_csv_export_returns_bytes(self) -> None:
        """export_csv() returns bytes."""
        records = [_make_record()]
        result = await EXPORTER.export_csv(records)
        assert isinstance(result, bytes)

    @pytest.mark.asyncio
    async def test_csv_export_starts_with_utf8_bom(self) -> None:
        """export_csv() output starts with the UTF-8 BOM so Excel detects Danish chars."""
        records = [_make_record()]
        result = await EXPORTER.export_csv(records)
        # UTF-8 BOM is EF BB BF
        assert result[:3] == b"\xef\xbb\xbf"

    @pytest.mark.asyncio
    async def test_csv_export_header_contains_all_flat_columns(self) -> None:
        """The CSV header row contains the human-readable label for every column
        in _FLAT_COLUMNS.

        Phase A refactoring: the CSV exporter now writes human-readable labels
        from ``_COLUMN_HEADERS`` (e.g. "Author" instead of
        ``author_display_name``) so that exported files are immediately
        legible to non-technical research users.  The test now checks for the
        mapped label, not the raw snake_case column name.
        """
        records = [_make_record()]
        result = await EXPORTER.export_csv(records)
        text = result.decode("utf-8-sig")
        header_line = text.splitlines()[0]
        for col in _FLAT_COLUMNS:
            expected_header = _COLUMN_HEADERS.get(col, col)
            assert expected_header in header_line, (
                f"Human-readable header {expected_header!r} (for column {col!r}) "
                f"missing from CSV header line"
            )

    @pytest.mark.asyncio
    async def test_csv_export_row_count_matches_records(self) -> None:
        """CSV has exactly N data rows for N input records (plus 1 header row)."""
        records = [_make_record(author_id=f"author-{i}") for i in range(5)]
        result = await EXPORTER.export_csv(records)
        lines = [ln for ln in result.decode("utf-8-sig").splitlines() if ln.strip()]
        # 1 header + 5 data rows
        assert len(lines) == 6

    @pytest.mark.asyncio
    async def test_csv_export_preserves_danish_characters(self) -> None:
        """Danish characters æ, ø, å survive CSV export without corruption."""
        records = [_make_record(text=DANISH_TEXT, display_name=DANISH_AUTHOR)]
        result = await EXPORTER.export_csv(records)
        decoded = result.decode("utf-8-sig")
        assert "Grøn" in decoded
        assert "velfærdsstaten" in decoded
        assert "Ålborg" in decoded
        assert "Søren" in decoded
        assert "Ærlighed" in decoded

    @pytest.mark.asyncio
    async def test_csv_export_empty_records_returns_header_only(self) -> None:
        """export_csv([]) returns bytes with a header row and no data rows.

        The header uses the human-readable label from _COLUMN_HEADERS, so
        check for "Platform" (the mapped label) rather than the snake_case
        key "platform".
        """
        result = await EXPORTER.export_csv([])
        text = result.decode("utf-8-sig")
        lines = [ln for ln in text.splitlines() if ln.strip()]
        assert len(lines) == 1
        assert _COLUMN_HEADERS.get("platform", "platform") in lines[0]

    @pytest.mark.asyncio
    async def test_csv_export_with_metadata_appends_raw_metadata_column(self) -> None:
        """include_metadata=True adds a raw_metadata column as JSON string.

        The column header uses the human-readable label from _COLUMN_HEADERS:
        "Raw Metadata (JSON)" rather than the snake_case key "raw_metadata".
        """
        meta = {"source": "bluesky", "raw_id": "abc123"}
        records = [_make_record(raw_metadata=meta)]
        result = await EXPORTER.export_csv(records, include_metadata=True)
        text = result.decode("utf-8-sig")
        # Human-readable header label (from _COLUMN_HEADERS) must appear in header row.
        assert _COLUMN_HEADERS.get("raw_metadata", "raw_metadata") in text.splitlines()[0]
        # The raw_metadata JSON should appear in the data row
        assert "bluesky" in text

    @pytest.mark.asyncio
    async def test_csv_export_list_field_joined_with_pipe(self) -> None:
        """search_terms_matched list is joined with ' | ' in CSV output."""
        records = [_make_record(terms=["klimaforandringer", "grøn omstilling"])]
        result = await EXPORTER.export_csv(records)
        text = result.decode("utf-8-sig")
        assert "klimaforandringer | grøn omstilling" in text


# ---------------------------------------------------------------------------
# XLSX export
# ---------------------------------------------------------------------------


class TestExportXlsx:
    @pytest.mark.asyncio
    async def test_xlsx_export_returns_bytes(self) -> None:
        """export_xlsx() returns non-empty bytes."""
        records = [_make_record()]
        result = await EXPORTER.export_xlsx(records)
        assert isinstance(result, bytes)
        assert len(result) > 0

    @pytest.mark.asyncio
    async def test_xlsx_export_produces_valid_xlsx_file(self, tmp_path) -> None:
        """export_xlsx() output is a valid XLSX file that openpyxl can load."""
        import openpyxl

        records = [_make_record()]
        xlsx_bytes = await EXPORTER.export_xlsx(records)
        path = tmp_path / "export.xlsx"
        path.write_bytes(xlsx_bytes)
        wb = openpyxl.load_workbook(path)
        assert wb is not None

    @pytest.mark.asyncio
    async def test_xlsx_export_uses_default_sheet_name(self, tmp_path) -> None:
        """Default sheet name is 'Content'."""
        import openpyxl

        records = [_make_record()]
        xlsx_bytes = await EXPORTER.export_xlsx(records)
        path = tmp_path / "export.xlsx"
        path.write_bytes(xlsx_bytes)
        wb = openpyxl.load_workbook(path)
        assert "Content" in wb.sheetnames

    @pytest.mark.asyncio
    async def test_xlsx_export_custom_sheet_name(self, tmp_path) -> None:
        """A custom sheet_name is used as the worksheet tab name."""
        import openpyxl

        records = [_make_record()]
        xlsx_bytes = await EXPORTER.export_xlsx(records, sheet_name="Klimadata")
        path = tmp_path / "export.xlsx"
        path.write_bytes(xlsx_bytes)
        wb = openpyxl.load_workbook(path)
        assert "Klimadata" in wb.sheetnames

    @pytest.mark.asyncio
    async def test_xlsx_export_preserves_danish_characters(self, tmp_path) -> None:
        """Danish characters æ, ø, å are preserved through XLSX export."""
        import openpyxl

        records = [_make_record(text=DANISH_TEXT, display_name=DANISH_AUTHOR)]
        xlsx_bytes = await EXPORTER.export_xlsx(records)
        path = tmp_path / "export.xlsx"
        path.write_bytes(xlsx_bytes)
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        # Collect all cell values into a single string for assertion
        all_values = " ".join(
            str(cell.value) for row in ws.iter_rows() for cell in row if cell.value is not None
        )
        assert "Grøn" in all_values
        assert "velfærdsstaten" in all_values
        assert "Søren" in all_values

    @pytest.mark.asyncio
    async def test_xlsx_export_empty_records_has_header_row_only(self, tmp_path) -> None:
        """export_xlsx([]) produces a workbook with one header row, no data rows."""
        import openpyxl

        xlsx_bytes = await EXPORTER.export_xlsx([])
        path = tmp_path / "export.xlsx"
        path.write_bytes(xlsx_bytes)
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        # Row 1 is the header; row 2 should not exist (max_row == 1)
        assert ws.max_row == 1

    @pytest.mark.asyncio
    async def test_xlsx_export_correct_row_count(self, tmp_path) -> None:
        """XLSX has exactly N+1 rows (1 header + N data rows)."""
        import openpyxl

        n = 7
        records = [_make_record(author_id=f"author-{i}") for i in range(n)]
        xlsx_bytes = await EXPORTER.export_xlsx(records)
        path = tmp_path / "export.xlsx"
        path.write_bytes(xlsx_bytes)
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        assert ws.max_row == n + 1


# ---------------------------------------------------------------------------
# JSON (NDJSON) export
# ---------------------------------------------------------------------------


class TestExportJson:
    @pytest.mark.asyncio
    async def test_json_export_returns_bytes(self) -> None:
        """export_json() returns bytes."""
        records = [_make_record()]
        result = await EXPORTER.export_json(records)
        assert isinstance(result, bytes)

    @pytest.mark.asyncio
    async def test_json_export_each_line_is_valid_json(self) -> None:
        """Each line in the NDJSON output is a valid JSON object."""
        records = [_make_record(author_id=f"author-{i}") for i in range(3)]
        result = await EXPORTER.export_json(records)
        lines = result.decode("utf-8").strip().split("\n")
        assert len(lines) == 3
        for line in lines:
            parsed = json.loads(line)
            assert isinstance(parsed, dict)

    @pytest.mark.asyncio
    async def test_json_export_all_records_present(self) -> None:
        """NDJSON output contains one JSON object per input record."""
        n = 4
        records = [_make_record(author_id=f"a-{i}") for i in range(n)]
        result = await EXPORTER.export_json(records)
        lines = result.decode("utf-8").strip().split("\n")
        assert len(lines) == n

    @pytest.mark.asyncio
    async def test_json_export_raw_metadata_intact(self) -> None:
        """raw_metadata dict is preserved exactly in NDJSON output."""
        meta = {"api_version": "1.0", "lang": "da", "score": 42}
        records = [_make_record(raw_metadata=meta)]
        result = await EXPORTER.export_json(records)
        parsed = json.loads(result.decode("utf-8").strip())
        assert parsed["raw_metadata"] == meta

    @pytest.mark.asyncio
    async def test_json_export_preserves_danish_characters(self) -> None:
        """Danish characters æ, ø, å are preserved in NDJSON output."""
        records = [_make_record(text=DANISH_TEXT, display_name=DANISH_AUTHOR)]
        result = await EXPORTER.export_json(records)
        decoded = result.decode("utf-8")
        assert "Grøn" in decoded
        assert "velfærdsstaten" in decoded
        assert "Søren" in decoded
        assert "Ærlighed" in decoded

    @pytest.mark.asyncio
    async def test_json_export_empty_records_returns_empty_bytes(self) -> None:
        """export_json([]) returns empty bytes (no lines, not a JSON array)."""
        result = await EXPORTER.export_json([])
        assert result == b""

    @pytest.mark.asyncio
    async def test_json_export_uses_ensure_ascii_false(self) -> None:
        """JSON output does not escape non-ASCII chars — native UTF-8 encoding."""
        records = [_make_record(text="æ ø å")]
        result = await EXPORTER.export_json(records)
        # If ensure_ascii=True were used, æ would be \\u00e6
        assert b"\\u00e6" not in result
        assert "æ".encode("utf-8") in result


# ---------------------------------------------------------------------------
# GEXF — actor co-occurrence (DQ-02 regression)
# ---------------------------------------------------------------------------


class TestExportGexfActor:
    """Tests for the actor co-occurrence GEXF builder.

    Critical regression: edge weight must equal the count of shared search terms,
    NOT the count of shared collection runs (the pre-DQ-02 bug).
    """

    @pytest.mark.asyncio
    async def test_gexf_actor_returns_bytes(self) -> None:
        """export_gexf() with network_type='actor' returns bytes."""
        records = [_make_record()]
        result = await EXPORTER.export_gexf(records, network_type="actor")
        assert isinstance(result, bytes)
        assert len(result) > 0

    @pytest.mark.asyncio
    async def test_gexf_actor_is_valid_xml(self) -> None:
        """The actor GEXF output parses as well-formed XML."""
        records = [
            _make_record(author_id="author-a"),
            _make_record(author_id="author-b"),
        ]
        result = await EXPORTER.export_gexf(records, network_type="actor")
        root = ET.fromstring(result.split(b"\n", 1)[1])  # skip XML declaration
        assert root.tag.endswith("gexf")

    @pytest.mark.asyncio
    async def test_gexf_actor_two_authors_one_term_produces_one_edge(self) -> None:
        """Two authors both matching the same term produce exactly one edge."""
        records = [
            _make_record(author_id="author-a", terms=["klimaforandringer"]),
            _make_record(author_id="author-b", terms=["klimaforandringer"]),
        ]
        result = await EXPORTER.export_gexf(records, network_type="actor")
        xml_body = result.split(b"\n", 1)[1]
        root = ET.fromstring(xml_body)
        ns = "http://gexf.net/1.3"
        edges = root.findall(f".//{{{ns}}}edge")
        assert len(edges) == 1

    @pytest.mark.asyncio
    async def test_gexf_actor_edge_weight_equals_shared_term_count(self) -> None:
        """DQ-02 regression: edge weight = number of distinct shared search terms.

        Author A and B both match 'klimaforandringer' and 'grøn omstilling'.
        Edge weight must be 2 (two shared terms), not 1 (one shared run).
        """
        records = [
            _make_record(
                author_id="author-a",
                terms=["klimaforandringer", "grøn omstilling"],
                run_id="run-001",
            ),
            _make_record(
                author_id="author-b",
                terms=["klimaforandringer", "grøn omstilling"],
                run_id="run-001",
            ),
        ]
        result = await EXPORTER.export_gexf(records, network_type="actor")
        xml_body = result.split(b"\n", 1)[1]
        root = ET.fromstring(xml_body)
        ns = "http://gexf.net/1.3"
        edges = root.findall(f".//{{{ns}}}edge")
        assert len(edges) == 1
        # Weight attribute must be 2 (shared term count), not 1 (shared run count)
        assert edges[0].get("weight") == "2"

    @pytest.mark.asyncio
    async def test_gexf_actor_grouping_by_term_not_by_run_id(self) -> None:
        """DQ-02 regression: two authors sharing only one term produce weight=1.

        Even if both records belong to the same run, weight must reflect
        the number of shared TERMS, not the number of shared collection runs.
        """
        records = [
            _make_record(
                author_id="author-a",
                terms=["klimaforandringer"],
                run_id="run-001",
            ),
            _make_record(
                author_id="author-b",
                terms=["klimaforandringer"],
                run_id="run-001",
            ),
            # Author C shares a different term with A — produces a separate edge
            _make_record(
                author_id="author-c",
                terms=["velfærdsstat"],
                run_id="run-001",
            ),
            _make_record(
                author_id="author-a",
                terms=["velfærdsstat"],
                run_id="run-001",
            ),
        ]
        result = await EXPORTER.export_gexf(records, network_type="actor")
        xml_body = result.split(b"\n", 1)[1]
        root = ET.fromstring(xml_body)
        ns = "http://gexf.net/1.3"
        edges = root.findall(f".//{{{ns}}}edge")

        # Build edge map by (source, target) pair
        edge_map: dict[tuple[str, str], str] = {}
        for e in edges:
            src = e.get("source", "")
            tgt = e.get("target", "")
            key = (min(src, tgt), max(src, tgt))
            edge_map[key] = e.get("weight", "")

        # A-B edge: only "klimaforandringer" is shared → weight = 1
        a_b_key = ("author-a", "author-b")
        assert a_b_key in edge_map, "Edge between author-a and author-b not found"
        assert edge_map[a_b_key] == "1", (
            f"Expected weight=1 for A-B (one shared term), got {edge_map[a_b_key]}"
        )

        # A-C edge: only "velfærdsstat" is shared → weight = 1
        a_c_key = ("author-a", "author-c")
        assert a_c_key in edge_map, "Edge between author-a and author-c not found"
        assert edge_map[a_c_key] == "1"

    @pytest.mark.asyncio
    async def test_gexf_actor_no_self_edges(self) -> None:
        """Actor GEXF contains no self-edges (source == target)."""
        records = [
            _make_record(author_id="author-a", terms=["term1", "term2"]),
            _make_record(author_id="author-b", terms=["term1"]),
        ]
        result = await EXPORTER.export_gexf(records, network_type="actor")
        xml_body = result.split(b"\n", 1)[1]
        root = ET.fromstring(xml_body)
        ns = "http://gexf.net/1.3"
        edges = root.findall(f".//{{{ns}}}edge")
        for edge in edges:
            assert edge.get("source") != edge.get("target"), "Self-edge found in GEXF output"

    @pytest.mark.asyncio
    async def test_gexf_actor_no_duplicate_nodes(self) -> None:
        """Actor GEXF nodes are unique — no actor appears twice."""
        # Author A appears in two records
        records = [
            _make_record(author_id="author-a", terms=["term1"]),
            _make_record(author_id="author-a", terms=["term2"]),
            _make_record(author_id="author-b", terms=["term1"]),
        ]
        result = await EXPORTER.export_gexf(records, network_type="actor")
        xml_body = result.split(b"\n", 1)[1]
        root = ET.fromstring(xml_body)
        ns = "http://gexf.net/1.3"
        node_ids = [node.get("id") for node in root.findall(f".//{{{ns}}}node")]
        assert len(node_ids) == len(set(node_ids)), "Duplicate node IDs found in GEXF"

    @pytest.mark.asyncio
    async def test_gexf_actor_records_without_author_id_skipped(self) -> None:
        """Records with pseudonymized_author_id=None are excluded from the graph."""
        records = [
            _make_record(author_id="author-a", terms=["term1"]),
            {  # Record with no author — should be silently skipped
                "pseudonymized_author_id": None,
                "search_terms_matched": ["term1"],
                "author_display_name": "Anonymous",
            },
        ]
        result = await EXPORTER.export_gexf(records, network_type="actor")
        xml_body = result.split(b"\n", 1)[1]
        root = ET.fromstring(xml_body)
        ns = "http://gexf.net/1.3"
        node_ids = [node.get("id") for node in root.findall(f".//{{{ns}}}node")]
        # Only author-a should appear; the anonymous record must not create a node
        assert "author-a" in node_ids
        assert None not in node_ids

    @pytest.mark.asyncio
    async def test_gexf_actor_empty_records_returns_valid_empty_gexf(self) -> None:
        """export_gexf([]) for actor type returns valid GEXF with empty nodes/edges."""
        result = await EXPORTER.export_gexf([], network_type="actor")
        xml_body = result.split(b"\n", 1)[1]
        root = ET.fromstring(xml_body)
        ns = "http://gexf.net/1.3"
        nodes = root.findall(f".//{{{ns}}}node")
        edges = root.findall(f".//{{{ns}}}edge")
        assert nodes == []
        assert edges == []

    @pytest.mark.asyncio
    async def test_gexf_actor_authors_sharing_no_terms_produce_no_edge(self) -> None:
        """Two authors with completely different terms have no edge in the network."""
        records = [
            _make_record(author_id="author-a", terms=["klimaforandringer"]),
            _make_record(author_id="author-b", terms=["velfærdsstat"]),
        ]
        result = await EXPORTER.export_gexf(records, network_type="actor")
        xml_body = result.split(b"\n", 1)[1]
        root = ET.fromstring(xml_body)
        ns = "http://gexf.net/1.3"
        edges = root.findall(f".//{{{ns}}}edge")
        assert edges == []

    @pytest.mark.asyncio
    async def test_gexf_actor_duplicate_content_records_not_double_counted(self) -> None:
        """Duplicate records for the same author do not artificially inflate edge weight.

        Author A has two records both matching the same term. Author B has one record.
        The edge A-B weight must be 1 (one shared term), not 2 (two records).
        """
        records = [
            # Author A appears twice with the same term
            _make_record(author_id="author-a", terms=["klimaforandringer"]),
            _make_record(author_id="author-a", terms=["klimaforandringer"]),
            # Author B appears once with the same term
            _make_record(author_id="author-b", terms=["klimaforandringer"]),
        ]
        result = await EXPORTER.export_gexf(records, network_type="actor")
        xml_body = result.split(b"\n", 1)[1]
        root = ET.fromstring(xml_body)
        ns = "http://gexf.net/1.3"
        edges = root.findall(f".//{{{ns}}}edge")
        assert len(edges) == 1
        # Weight = number of distinct shared TERMS, not records → must be 1
        assert edges[0].get("weight") == "1"

    @pytest.mark.asyncio
    async def test_gexf_actor_shared_terms_attribute_contains_intersection(self) -> None:
        """The shared_terms edge attribute contains only the terms both authors share."""
        records = [
            _make_record(
                author_id="author-a",
                terms=["klimaforandringer", "grøn omstilling", "demokrati"],
            ),
            _make_record(
                author_id="author-b",
                terms=["klimaforandringer", "grøn omstilling"],
            ),
        ]
        result = await EXPORTER.export_gexf(records, network_type="actor")
        xml_body = result.split(b"\n", 1)[1]
        root = ET.fromstring(xml_body)
        ns = "http://gexf.net/1.3"
        edges = root.findall(f".//{{{ns}}}edge")
        assert len(edges) == 1

        # The attvalue for "shared_terms" (for="1") should contain only the shared terms
        attvalues = edges[0].findall(f".//{{{ns}}}attvalue")
        shared_terms_val = None
        for av in attvalues:
            if av.get("for") == "1":
                shared_terms_val = av.get("value", "")
                break
        assert shared_terms_val is not None, "shared_terms attvalue not found"
        # Both shared terms should be present
        assert "klimaforandringer" in shared_terms_val
        assert "grøn omstilling" in shared_terms_val
        # The non-shared term must not appear
        assert "demokrati" not in shared_terms_val

    @pytest.mark.asyncio
    async def test_gexf_actor_unknown_network_type_raises_value_error(self) -> None:
        """export_gexf() raises ValueError for an unknown network_type."""
        with pytest.raises(ValueError, match="Unknown network_type"):
            await EXPORTER.export_gexf([], network_type="bipartite_extended")


# ---------------------------------------------------------------------------
# GEXF — term co-occurrence
# ---------------------------------------------------------------------------


class TestExportGexfTerm:
    @pytest.mark.asyncio
    async def test_gexf_term_two_terms_in_same_record_produce_edge(self) -> None:
        """Two terms co-occurring in the same record produce an edge."""
        records = [
            _make_record(terms=["klimaforandringer", "grøn omstilling"]),
        ]
        result = await EXPORTER.export_gexf(records, network_type="term")
        xml_body = result.split(b"\n", 1)[1]
        root = ET.fromstring(xml_body)
        ns = "http://gexf.net/1.3"
        edges = root.findall(f".//{{{ns}}}edge")
        assert len(edges) == 1

    @pytest.mark.asyncio
    async def test_gexf_term_single_term_record_produces_no_edges(self) -> None:
        """A record with only one term produces no edges (no co-occurrence)."""
        records = [
            _make_record(terms=["klimaforandringer"]),
        ]
        result = await EXPORTER.export_gexf(records, network_type="term")
        xml_body = result.split(b"\n", 1)[1]
        root = ET.fromstring(xml_body)
        ns = "http://gexf.net/1.3"
        edges = root.findall(f".//{{{ns}}}edge")
        nodes = root.findall(f".//{{{ns}}}node")
        assert edges == []
        # Single-term nodes should also be absent (no edges → no edge_terms)
        assert nodes == []

    @pytest.mark.asyncio
    async def test_gexf_term_node_type_attribute_is_term(self) -> None:
        """All nodes in a term co-occurrence GEXF have type='term'."""
        records = [
            _make_record(terms=["klimaforandringer", "grøn omstilling"]),
        ]
        result = await EXPORTER.export_gexf(records, network_type="term")
        xml_body = result.split(b"\n", 1)[1]
        root = ET.fromstring(xml_body)
        ns = "http://gexf.net/1.3"
        # Check only node attvalues (not edge attvalues which also use for="0")
        nodes_el = root.find(f".//{{{ns}}}nodes")
        assert nodes_el is not None
        for node in nodes_el.findall(f"{{{ns}}}node"):
            attvalues = node.findall(f".//{{{ns}}}attvalue[@for='0']")
            for av in attvalues:
                assert av.get("value") == "term"

    @pytest.mark.asyncio
    async def test_gexf_term_empty_records_returns_valid_empty_gexf(self) -> None:
        """export_gexf([]) for term type returns valid GEXF with no nodes/edges."""
        result = await EXPORTER.export_gexf([], network_type="term")
        xml_body = result.split(b"\n", 1)[1]
        root = ET.fromstring(xml_body)
        ns = "http://gexf.net/1.3"
        assert root.findall(f".//{{{ns}}}node") == []
        assert root.findall(f".//{{{ns}}}edge") == []


# ---------------------------------------------------------------------------
# GEXF — bipartite actor-term
# ---------------------------------------------------------------------------


class TestExportGexfBipartite:
    @pytest.mark.asyncio
    async def test_gexf_bipartite_returns_bytes(self) -> None:
        """export_gexf() with network_type='bipartite' returns bytes."""
        records = [_make_record(author_id="author-a", terms=["klimaforandringer"])]
        result = await EXPORTER.export_gexf(records, network_type="bipartite")
        assert isinstance(result, bytes)

    @pytest.mark.asyncio
    async def test_gexf_bipartite_term_node_ids_have_term_prefix(self) -> None:
        """Term nodes in bipartite GEXF have IDs prefixed with 'term:'."""
        records = [_make_record(author_id="author-a", terms=["klimaforandringer"])]
        result = await EXPORTER.export_gexf(records, network_type="bipartite")
        xml_body = result.split(b"\n", 1)[1]
        root = ET.fromstring(xml_body)
        ns = "http://gexf.net/1.3"
        nodes = root.findall(f".//{{{ns}}}node")
        term_nodes = [n for n in nodes if (n.get("id") or "").startswith("term:")]
        assert len(term_nodes) >= 1
        assert any(n.get("id") == "term:klimaforandringer" for n in term_nodes)

    @pytest.mark.asyncio
    async def test_gexf_bipartite_actor_and_term_node_types(self) -> None:
        """Bipartite GEXF contains both 'actor' and 'term' type nodes."""
        records = [_make_record(author_id="author-a", terms=["klimaforandringer"])]
        result = await EXPORTER.export_gexf(records, network_type="bipartite")
        xml_body = result.split(b"\n", 1)[1]
        root = ET.fromstring(xml_body)
        ns = "http://gexf.net/1.3"
        type_values = {
            av.get("value")
            for av in root.findall(f".//{{{ns}}}attvalue[@for='0']")
        }
        assert "actor" in type_values
        assert "term" in type_values

    @pytest.mark.asyncio
    async def test_gexf_bipartite_empty_records_returns_valid_empty_gexf(self) -> None:
        """export_gexf([]) for bipartite type returns valid GEXF with no content."""
        result = await EXPORTER.export_gexf([], network_type="bipartite")
        xml_body = result.split(b"\n", 1)[1]
        root = ET.fromstring(xml_body)
        ns = "http://gexf.net/1.3"
        assert root.findall(f".//{{{ns}}}node") == []
        assert root.findall(f".//{{{ns}}}edge") == []


# ---------------------------------------------------------------------------
# H-02: RIS export
# ---------------------------------------------------------------------------


class TestExportRis:
    def test_ris_returns_bytes(self) -> None:
        """export_ris() returns bytes."""
        records = [_make_record()]
        result = EXPORTER.export_ris(records)
        assert isinstance(result, bytes)

    def test_ris_each_record_ends_with_er_tag(self) -> None:
        """Three records produce exactly three 'ER  - ' end-of-record markers."""
        records = [_make_record(author_id=f"a-{i}") for i in range(3)]
        result = EXPORTER.export_ris(records)
        decoded = result.decode("utf-8")
        assert decoded.count("ER  - ") == 3

    def test_ris_starts_with_ty_elec(self) -> None:
        """RIS output begins with 'TY  - ELEC' (Electronic Source type tag)."""
        records = [_make_record()]
        result = EXPORTER.export_ris(records)
        decoded = result.decode("utf-8")
        assert decoded.startswith("TY  - ELEC")

    def test_ris_title_tag_present_when_title_set(self) -> None:
        """TI tag is present when the record has a title field."""
        record = _make_record()
        record["title"] = "Klimaforandringer i Arktis"
        result = EXPORTER.export_ris([record])
        decoded = result.decode("utf-8")
        assert "TI  - Klimaforandringer i Arktis" in decoded

    def test_ris_author_tag_present(self) -> None:
        """AU  -  tag is present in the output when author_display_name is set."""
        record = _make_record(display_name="Søren Øberg")
        result = EXPORTER.export_ris([record])
        decoded = result.decode("utf-8")
        assert "AU  - " in decoded

    def test_ris_url_tag_present(self) -> None:
        """UR  -  tag is present in the output when url is set."""
        record = _make_record(author_id="author-url-test")
        result = EXPORTER.export_ris([record])
        decoded = result.decode("utf-8")
        assert "UR  - " in decoded

    def test_ris_preserves_danish_characters(self) -> None:
        """Danish characters æ, ø, å survive RIS export without corruption."""
        record = _make_record(display_name=DANISH_AUTHOR, text=DANISH_TEXT)
        result = EXPORTER.export_ris([record])
        decoded = result.decode("utf-8")
        assert "Grøn" in decoded
        assert "Søren" in decoded

    def test_ris_empty_records_returns_empty_bytes(self) -> None:
        """export_ris([]) returns empty bytes (no content, not an error)."""
        result = EXPORTER.export_ris([])
        assert result == b""


# ---------------------------------------------------------------------------
# H-02: BibTeX export
# ---------------------------------------------------------------------------


class TestExportBibTeX:
    def test_bibtex_returns_bytes(self) -> None:
        """export_bibtex() returns bytes."""
        records = [_make_record()]
        result = EXPORTER.export_bibtex(records)
        assert isinstance(result, bytes)

    def test_bibtex_entry_type_is_misc(self) -> None:
        """Every BibTeX entry uses the @misc type."""
        records = [_make_record()]
        result = EXPORTER.export_bibtex(records)
        decoded = result.decode("utf-8")
        assert "@misc{" in decoded

    def test_bibtex_entry_key_derived_from_content_hash(self) -> None:
        """The entry key contains the first 8 characters of the content_hash."""
        record = _make_record()
        record["content_hash"] = "abcdef1234567890" + "a" * 48
        result = EXPORTER.export_bibtex([record])
        decoded = result.decode("utf-8")
        assert "record_abcdef12" in decoded

    def test_bibtex_curly_braces_escaped_in_title(self) -> None:
        """Curly braces in the title are escaped as \\{ and \\}."""
        record = _make_record()
        record["title"] = "Title with {braces}"
        result = EXPORTER.export_bibtex([record])
        decoded = result.decode("utf-8")
        assert r"\{braces\}" in decoded

    def test_bibtex_preserves_danish_characters(self) -> None:
        """Danish characters æ, ø, å survive BibTeX export without corruption."""
        record = _make_record(display_name=DANISH_AUTHOR, text=DANISH_TEXT)
        result = EXPORTER.export_bibtex([record])
        decoded = result.decode("utf-8")
        assert "Grøn" in decoded
        assert "Søren" in decoded

    def test_bibtex_unique_keys_for_distinct_hashes(self) -> None:
        """Two records with different content_hash values produce different entry keys."""
        import re

        record_a = _make_record(author_id="a1")
        record_a["content_hash"] = "aaaa" * 16
        record_b = _make_record(author_id="b2")
        record_b["content_hash"] = "bbbb" * 16

        result = EXPORTER.export_bibtex([record_a, record_b])
        decoded = result.decode("utf-8")

        # Parse all @misc entry keys from the output.
        keys = re.findall(r"@misc\{(record_\w+),", decoded)
        assert len(keys) == 2, f"Expected 2 entry keys, got: {keys}"
        assert len(set(keys)) == 2, f"Duplicate entry keys found: {keys}"
