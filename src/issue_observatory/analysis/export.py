"""Content export utilities for the Issue Observatory.

Exports ``content_records`` data to CSV, XLSX, JSON (NDJSON), Parquet, and
GEXF formats.  All format methods are async and return raw bytes ready to be
streamed as an HTTP response or uploaded to object storage.

The ``ContentExporter`` class is format-agnostic:

- For CSV, XLSX, JSON, and Parquet exports, callers pass in a list of plain
  dicts (typically produced by a SQLAlchemy ``mappings()`` query result or a
  list comprehension over ORM rows).
- For GEXF exports, callers pass in a **graph dict** produced by one of the
  network analysis functions in :mod:`issue_observatory.analysis.network`
  (e.g. :func:`~issue_observatory.analysis.network.get_actor_co_occurrence`).
  The graph dict has the shape ``{"nodes": [...], "edges": [...]}``.  This
  separation ensures that duplicate exclusion and scoping filters are applied
  at the database layer (inside ``network.py``), not reconstructed in the
  exporter.

The exporter does NOT perform database queries itself — that responsibility
stays with the route handler or Celery task.

Optional dependencies:
    openpyxl  — required for ``export_xlsx``; raises ImportError if missing.
    pyarrow   — required for ``export_parquet``; raises ImportError if missing.

Both are listed as main project dependencies in pyproject.toml (Phase 3 core
feature), so ImportError should never occur in production.  The guard exists
for environments that install a minimal subset of dependencies.

Owned by the DB Engineer.
"""

from __future__ import annotations

import io
import json
import uuid as uuid_mod
import xml.etree.ElementTree as ET
from collections import defaultdict
from datetime import datetime, timezone
from typing import Any

import structlog

logger = structlog.get_logger(__name__)

# ---------------------------------------------------------------------------
# Column definitions
# ---------------------------------------------------------------------------

#: Ordered list of flat columns written by CSV and XLSX exporters.
_FLAT_COLUMNS: list[str] = [
    "platform",
    "arena",
    "content_type",
    "title",
    "text_content",
    "url",
    "author_display_name",
    "pseudonymized_author_id",
    "published_at",
    "views_count",
    "likes_count",
    "shares_count",
    "comments_count",
    "engagement_score",
    "language",
    "collection_tier",
    "search_terms_matched",
    "content_hash",
    "collection_run_id",
    "query_design_id",
]

#: Human-readable column header labels for CSV and XLSX output.
#:
#: Keys are the snake_case column names from ``_FLAT_COLUMNS``; values are
#: the labels written to the header row so that exported files are immediately
#: legible to non-technical research users.
_COLUMN_HEADERS: dict[str, str] = {
    "platform": "Platform",
    "arena": "Arena",
    "content_type": "Content Type",
    "title": "Title",
    "text_content": "Text Content",
    "url": "URL",
    "author_display_name": "Author",
    "pseudonymized_author_id": "Author ID (Pseudonymized)",
    "published_at": "Published At",
    "views_count": "Views",
    "likes_count": "Likes",
    "shares_count": "Shares",
    "comments_count": "Comments",
    "engagement_score": "Engagement Score",
    "language": "Language",
    "collection_tier": "Collection Tier",
    "search_terms_matched": "Matched Search Terms",
    "content_hash": "Content Hash",
    "collection_run_id": "Collection Run ID",
    "query_design_id": "Query Design ID",
    # raw_metadata is an optional trailing column in CSV only (include_metadata=True).
    "raw_metadata": "Raw Metadata (JSON)",
}


def _safe_str(value: Any) -> str:  # noqa: ANN401
    """Coerce a record value to a safe UTF-8 string for CSV/XLSX output.

    Lists (e.g. ``search_terms_matched``) are joined with ``|`` so that each
    row stays a single cell.  None becomes an empty string.

    Args:
        value: Any Python value from a deserialized content record dict.

    Returns:
        A string representation safe for inclusion in a spreadsheet cell.
    """
    if value is None:
        return ""
    if isinstance(value, list):
        return " | ".join(str(v) for v in value)
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value)


class ContentExporter:
    """Export content records to various file formats.

    All public methods are ``async`` for consistency with the FastAPI/Celery
    calling context, even though the work is CPU-bound (I/O happens via in-memory
    buffers).  If profiling shows that large exports block the event loop, the
    CPU-intensive sections should be wrapped in ``asyncio.to_thread``.

    Typical usage in a route handler::

        exporter = ContentExporter()
        data = await exporter.export_csv(records, include_metadata=True)
        return Response(content=data, media_type="text/csv; charset=utf-8")
    """

    # ------------------------------------------------------------------
    # CSV
    # ------------------------------------------------------------------

    async def export_csv(
        self,
        records: list[dict[str, Any]],
        include_metadata: bool = False,
    ) -> bytes:
        """Export records as a UTF-8 CSV file.

        Columns written in order match ``_FLAT_COLUMNS``.  Header labels come
        from ``_COLUMN_HEADERS`` so the exported file uses human-readable
        column names (e.g. "Author" instead of ``author_display_name``).

        If ``include_metadata`` is True, a final column ``Raw Metadata (JSON)``
        contains the JSONB payload serialized as a JSON string.

        Args:
            records: List of content record dicts (keys match ORM column names).
            include_metadata: Whether to include the ``raw_metadata`` JSONB
                column as a trailing JSON string column.

        Returns:
            Raw UTF-8 encoded CSV bytes including the BOM marker so that
            Microsoft Excel opens it correctly without charset configuration.
        """
        import csv  # stdlib — always available

        # snake_case column list drives data extraction from the record dicts.
        columns = list(_FLAT_COLUMNS)
        if include_metadata:
            columns.append("raw_metadata")

        # Human-readable header labels (falls back to the snake_case name for
        # any column not present in _COLUMN_HEADERS, e.g. raw_metadata).
        headers = [_COLUMN_HEADERS.get(col, col) for col in columns]

        buf = io.StringIO()
        writer = csv.writer(buf, lineterminator="\r\n")
        writer.writerow(headers)

        for rec in records:
            row: list[str] = [_safe_str(rec.get(col)) for col in columns]
            if include_metadata:
                meta = rec.get("raw_metadata")
                row[-1] = json.dumps(meta, ensure_ascii=False) if meta else ""
            writer.writerow(row)

        # UTF-8 BOM so Excel auto-detects encoding for Danish characters (æøå).
        return "\ufeff".encode("utf-8") + buf.getvalue().encode("utf-8")

    # ------------------------------------------------------------------
    # XLSX
    # ------------------------------------------------------------------

    async def export_xlsx(
        self,
        records: list[dict[str, Any]],
        sheet_name: str = "Content",
    ) -> bytes:
        """Export records as an XLSX workbook (UTF-8 safe for Danish æøå).

        The header row is bold and frozen; columns are auto-sized to the
        wider of the header or the longest value in the first 100 rows.

        Args:
            records: List of content record dicts.
            sheet_name: Name of the worksheet tab (max 31 chars, enforced by
                openpyxl automatically).

        Returns:
            Raw XLSX bytes.

        Raises:
            ImportError: If ``openpyxl`` is not installed.
        """
        try:
            import openpyxl
            from openpyxl.styles import Font
        except ImportError as exc:
            raise ImportError(
                "openpyxl is required for XLSX export. "
                "Install it with: pip install 'openpyxl>=3.1,<4.0'"
            ) from exc

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]  # Excel limit

        # snake_case column list drives data extraction from the record dicts.
        columns = list(_FLAT_COLUMNS)
        # Human-readable header labels for the worksheet header row.
        headers = [_COLUMN_HEADERS.get(col, col) for col in columns]

        # Write header row using human-readable labels.
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        ws.freeze_panes = "A2"

        # Write data rows (read data by snake_case column name).
        for rec in records:
            row_values = []
            for col in columns:
                val = rec.get(col)
                if val is None:
                    row_values.append("")
                elif isinstance(val, list):
                    row_values.append(" | ".join(str(v) for v in val))
                elif isinstance(val, datetime):
                    # Keep as datetime so Excel formats it as a date cell.
                    row_values.append(val.replace(tzinfo=None))
                elif isinstance(val, uuid_mod.UUID):
                    row_values.append(str(val))
                else:
                    row_values.append(val)
            ws.append(row_values)

        # Auto-size columns based on header label width + first 100 data rows.
        for col_idx, header_label in enumerate(headers, start=1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            max_len = len(header_label)
            for row_idx in range(2, min(102, ws.max_row + 1)):
                cell_val = ws.cell(row=row_idx, column=col_idx).value
                if cell_val is not None:
                    max_len = max(max_len, len(str(cell_val)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 80)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ------------------------------------------------------------------
    # JSON / NDJSON
    # ------------------------------------------------------------------

    async def export_json(self, records: list[dict[str, Any]]) -> bytes:
        """Export records as newline-delimited JSON (NDJSON).

        NDJSON is preferred over a single JSON array for large datasets because
        it can be streamed line-by-line without loading the full payload into
        memory.  Each line is a complete JSON object terminated by ``\\n``.

        Datetime objects are serialized to ISO 8601 strings.  UUID objects are
        serialized to their string representation.

        Args:
            records: List of content record dicts.

        Returns:
            UTF-8 encoded NDJSON bytes.
        """
        import uuid

        def _default(obj: Any) -> str:  # noqa: ANN401
            if isinstance(obj, datetime):
                return obj.isoformat()
            if isinstance(obj, uuid.UUID):
                return str(obj)
            raise TypeError(f"Object of type {type(obj).__name__} is not JSON serializable")

        lines = [
            json.dumps(rec, ensure_ascii=False, default=_default) for rec in records
        ]
        return "\n".join(lines).encode("utf-8")

    # ------------------------------------------------------------------
    # Parquet
    # ------------------------------------------------------------------

    async def export_parquet(self, records: list[dict[str, Any]]) -> bytes:
        """Export records as a Parquet file using pyarrow.

        Schema is inferred from the ``_FLAT_COLUMNS`` list.  Columns are
        typed as ``string`` for text fields; ``int64`` for count fields;
        ``timestamp[us, tz=UTC]`` for temporal fields.  The ``raw_metadata``
        column is omitted from Parquet output because nested JSONB does not
        map cleanly to a columnar schema — callers who need it should use the
        JSON or CSV exporter.

        Args:
            records: List of content record dicts.

        Returns:
            Raw Parquet bytes.

        Raises:
            ImportError: If ``pyarrow`` is not installed.
        """
        try:
            import pyarrow as pa
            import pyarrow.parquet as pq
        except ImportError as exc:
            raise ImportError(
                "pyarrow is required for Parquet export. "
                "Install it with: pip install 'pyarrow>=15.0,<16.0'"
            ) from exc

        import uuid

        string_cols = [
            "platform", "arena", "content_type", "title", "text_content",
            "url", "author_display_name", "pseudonymized_author_id",
            "language", "collection_tier",
            "content_hash", "collection_run_id", "query_design_id",
        ]
        int_cols = ["views_count", "likes_count", "shares_count", "comments_count"]
        float_cols = ["engagement_score"]
        ts_cols = ["published_at"]

        column_data: dict[str, list[Any]] = {col: [] for col in _FLAT_COLUMNS}

        for rec in records:
            for col in _FLAT_COLUMNS:
                val = rec.get(col)
                if col in int_cols:
                    column_data[col].append(int(val) if val is not None else None)
                elif col in float_cols:
                    column_data[col].append(float(val) if val is not None else None)
                elif col in ts_cols:
                    if isinstance(val, datetime):
                        # Convert to UTC-aware, then to microseconds timestamp
                        if val.tzinfo is None:
                            val = val.replace(tzinfo=timezone.utc)
                        column_data[col].append(val)
                    else:
                        column_data[col].append(None)
                elif col == "search_terms_matched":
                    if isinstance(val, list):
                        column_data[col].append(val)
                    else:
                        column_data[col].append([])
                elif isinstance(val, uuid.UUID):
                    column_data[col].append(str(val))
                else:
                    column_data[col].append(str(val) if val is not None else None)

        schema_fields = []
        for col in _FLAT_COLUMNS:
            if col in int_cols:
                schema_fields.append(pa.field(col, pa.int64()))
            elif col in float_cols:
                schema_fields.append(pa.field(col, pa.float64()))
            elif col in ts_cols:
                schema_fields.append(pa.field(col, pa.timestamp("us", tz="UTC")))
            elif col == "search_terms_matched":
                schema_fields.append(pa.field(col, pa.list_(pa.string())))
            else:
                schema_fields.append(pa.field(col, pa.string()))

        schema = pa.schema(schema_fields)
        arrays = []
        for col in _FLAT_COLUMNS:
            if col in int_cols:
                arrays.append(pa.array(column_data[col], type=pa.int64()))
            elif col in float_cols:
                arrays.append(pa.array(column_data[col], type=pa.float64()))
            elif col in ts_cols:
                arrays.append(pa.array(column_data[col], type=pa.timestamp("us", tz="UTC")))
            elif col == "search_terms_matched":
                arrays.append(pa.array(column_data[col], type=pa.list_(pa.string())))
            else:
                arrays.append(pa.array(column_data[col], type=pa.string()))

        table = pa.table(dict(zip(_FLAT_COLUMNS, arrays, strict=True)), schema=schema)
        buf = io.BytesIO()
        pq.write_table(table, buf)
        return buf.getvalue()

    # ------------------------------------------------------------------
    # GEXF — shared serialization helper
    # ------------------------------------------------------------------

    @staticmethod
    def _serialize_gexf(gexf_root: ET.Element) -> bytes:
        """Serialize a GEXF ``Element`` tree to indented UTF-8 XML bytes.

        Prepends the XML declaration line so the output is a fully-formed
        XML document.

        Args:
            gexf_root: The ``<gexf>`` root ``Element``.

        Returns:
            UTF-8 encoded XML bytes.
        """
        tree = ET.ElementTree(gexf_root)
        ET.indent(tree, space="  ")
        buf = io.StringIO()
        tree.write(buf, encoding="unicode", xml_declaration=False)
        xml_str = '<?xml version="1.0" encoding="UTF-8"?>\n' + buf.getvalue()
        return xml_str.encode("utf-8")

    @staticmethod
    def _make_gexf_root(description: str) -> tuple[ET.Element, ET.Element]:
        """Create the outer ``<gexf>`` and ``<graph>`` elements with standard metadata.

        Args:
            description: Human-readable description written into ``<meta>``.

        Returns:
            A ``(gexf_root, graph_element)`` tuple ready for node/edge children.
        """
        today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        gexf = ET.Element(
            "gexf",
            {
                "xmlns": "http://gexf.net/1.3",
                "xmlns:xsi": "http://www.w3.org/2001/XMLSchema-instance",
                "xsi:schemaLocation": (
                    "http://gexf.net/1.3 http://gexf.net/1.3/gexf.xsd"
                ),
                "version": "1.3",
            },
        )
        meta = ET.SubElement(gexf, "meta", {"lastmodifieddate": today})
        ET.SubElement(meta, "creator").text = "Issue Observatory"
        ET.SubElement(meta, "description").text = description

        graph = ET.SubElement(
            gexf, "graph", {"mode": "static", "defaultedgetype": "undirected"}
        )
        return gexf, graph

    # ------------------------------------------------------------------
    # Graph builders: records → graph dict
    # ------------------------------------------------------------------

    @staticmethod
    def _records_to_actor_graph(records: list[dict[str, Any]]) -> dict[str, Any]:
        """Build an actor co-occurrence graph dict from content records.

        Two actors are connected when they share at least one search term.
        Edge weight = number of distinct shared terms.
        """
        from collections import defaultdict as _dd
        from itertools import combinations as _comb

        # author_id → {terms}, platform, post_count
        author_terms: dict[str, set[str]] = _dd(set)
        author_platform: dict[str, str] = {}
        author_post_count: dict[str, int] = _dd(int)

        for rec in records:
            aid = rec.get("pseudonymized_author_id")
            if not aid:
                continue
            terms = rec.get("search_terms_matched") or []
            for t in terms:
                author_terms[aid].add(t)
            author_platform[aid] = rec.get("platform") or ""
            author_post_count[aid] += 1

        nodes = [
            {
                "id": aid,
                "label": aid,
                "platform": author_platform.get(aid, ""),
                "post_count": author_post_count[aid],
                "degree": 0,
            }
            for aid in author_terms
        ]

        edges: list[dict[str, Any]] = []
        degree: dict[str, int] = _dd(int)
        for a, b in _comb(sorted(author_terms.keys()), 2):
            shared = author_terms[a] & author_terms[b]
            if shared:
                edges.append({
                    "source": a,
                    "target": b,
                    "weight": len(shared),
                    "shared_terms": list(shared),
                })
                degree[a] += 1
                degree[b] += 1

        for node in nodes:
            node["degree"] = degree[node["id"]]

        return {"nodes": nodes, "edges": edges}

    @staticmethod
    def _records_to_term_graph(records: list[dict[str, Any]]) -> dict[str, Any]:
        """Build a term co-occurrence graph dict from content records.

        Only terms that co-occur with at least one other term (i.e. have edges)
        are included as nodes.
        """
        from collections import defaultdict as _dd
        from itertools import combinations as _comb

        term_freq: dict[str, int] = _dd(int)
        edge_weight: dict[tuple[str, str], int] = _dd(int)

        for rec in records:
            terms = sorted(set(rec.get("search_terms_matched") or []))
            for t in terms:
                term_freq[t] += 1
            for a, b in _comb(terms, 2):
                edge_weight[(a, b)] += 1

        edges = [
            {"source": a, "target": b, "weight": w}
            for (a, b), w in edge_weight.items()
        ]
        # Only include terms that appear in at least one edge
        connected_terms = {e["source"] for e in edges} | {e["target"] for e in edges}
        degree: dict[str, int] = _dd(int)
        for e in edges:
            degree[e["source"]] += 1
            degree[e["target"]] += 1

        nodes = [
            {
                "id": t,
                "label": t,
                "type": "term",
                "frequency": term_freq[t],
                "degree": degree[t],
            }
            for t in connected_terms
        ]

        return {"nodes": nodes, "edges": edges}

    @staticmethod
    def _records_to_bipartite_graph(records: list[dict[str, Any]]) -> dict[str, Any]:
        """Build a bipartite actor-term graph dict from content records."""
        from collections import defaultdict as _dd

        actor_platform: dict[str, str] = {}
        edge_weight: dict[tuple[str, str], int] = _dd(int)

        for rec in records:
            aid = rec.get("pseudonymized_author_id")
            if not aid:
                continue
            actor_platform[aid] = rec.get("platform") or ""
            for t in rec.get("search_terms_matched") or []:
                edge_weight[(aid, t)] += 1

        actor_ids = set(actor_platform.keys())
        term_ids = {t for _, t in edge_weight}

        nodes = [
            {"id": aid, "label": aid, "type": "actor", "platform": actor_platform[aid]}
            for aid in actor_ids
        ] + [
            {"id": f"term:{t}", "label": t, "type": "term"}
            for t in term_ids
        ]
        edges = [
            {"source": aid, "target": f"term:{t}", "weight": w}
            for (aid, t), w in edge_weight.items()
        ]

        return {"nodes": nodes, "edges": edges}

    # ------------------------------------------------------------------
    # GEXF — actor co-occurrence
    # ------------------------------------------------------------------

    def _build_actor_gexf(self, graph: dict[str, Any]) -> bytes:
        """Serialize an actor co-occurrence graph dict to GEXF.

        Consumes the graph dict format returned by
        :func:`~issue_observatory.analysis.network.get_actor_co_occurrence`:

        - **Nodes**: ``id``, ``label``, ``platform``, ``post_count``,
          ``degree``
        - **Edges**: ``source``, ``target``, ``weight``

        The graph dict is produced by the database query in ``network.py``
        and therefore already excludes duplicate-flagged records via the
        shared filter builder.  This method only performs XML serialization.

        Args:
            graph: Graph dict ``{"nodes": [...], "edges": [...]}`` as returned
                by :func:`~issue_observatory.analysis.network.get_actor_co_occurrence`.

        Returns:
            UTF-8 encoded GEXF XML bytes.
        """
        gexf, graph_el = self._make_gexf_root("Actor co-occurrence network")

        node_attrs = ET.SubElement(graph_el, "attributes", {"class": "node"})
        ET.SubElement(node_attrs, "attribute", {"id": "0", "title": "platform", "type": "string"})
        ET.SubElement(node_attrs, "attribute", {"id": "1", "title": "post_count", "type": "integer"})
        ET.SubElement(node_attrs, "attribute", {"id": "2", "title": "degree", "type": "integer"})

        edge_attrs = ET.SubElement(graph_el, "attributes", {"class": "edge"})
        ET.SubElement(edge_attrs, "attribute", {"id": "0", "title": "weight", "type": "float"})
        ET.SubElement(edge_attrs, "attribute", {"id": "1", "title": "shared_terms", "type": "string"})

        nodes_el = ET.SubElement(graph_el, "nodes")
        for node in graph.get("nodes", []):
            node_id = str(node["id"])
            label = str(node.get("label") or node_id)
            n = ET.SubElement(nodes_el, "node", {"id": node_id, "label": label})
            attvals = ET.SubElement(n, "attvalues")
            ET.SubElement(attvals, "attvalue", {
                "for": "0", "value": str(node.get("platform") or ""),
            })
            ET.SubElement(attvals, "attvalue", {
                "for": "1", "value": str(node.get("post_count") or 0),
            })
            ET.SubElement(attvals, "attvalue", {
                "for": "2", "value": str(node.get("degree") or 0),
            })

        edges_el = ET.SubElement(graph_el, "edges")
        for edge_idx, edge in enumerate(graph.get("edges", [])):
            e = ET.SubElement(
                edges_el,
                "edge",
                {
                    "id": str(edge_idx),
                    "source": str(edge["source"]),
                    "target": str(edge["target"]),
                    "weight": str(edge.get("weight") or 1),
                },
            )
            attvals = ET.SubElement(e, "attvalues")
            ET.SubElement(attvals, "attvalue", {
                "for": "0", "value": str(edge.get("weight") or 1),
            })
            shared = edge.get("shared_terms") or []
            if isinstance(shared, list):
                shared_str = "|".join(sorted(shared))
            else:
                shared_str = str(shared)
            ET.SubElement(attvals, "attvalue", {
                "for": "1", "value": shared_str,
            })

        return self._serialize_gexf(gexf)

    # ------------------------------------------------------------------
    # GEXF — term co-occurrence
    # ------------------------------------------------------------------

    def _build_term_gexf(self, graph: dict[str, Any]) -> bytes:
        """Serialize a term co-occurrence graph dict to GEXF.

        Consumes the graph dict format returned by
        :func:`~issue_observatory.analysis.network.get_term_co_occurrence`:

        - **Nodes**: ``id`` (term string), ``label``, ``type`` = ``"term"``,
          ``frequency``, ``degree``
        - **Edges**: ``source``, ``target``, ``weight``

        The graph dict is produced by the database query in ``network.py``
        and therefore already excludes duplicate-flagged records.

        Args:
            graph: Graph dict ``{"nodes": [...], "edges": [...]}`` as returned
                by :func:`~issue_observatory.analysis.network.get_term_co_occurrence`.

        Returns:
            UTF-8 encoded GEXF XML bytes.
        """
        gexf, graph_el = self._make_gexf_root("Term co-occurrence network")

        node_attrs = ET.SubElement(graph_el, "attributes", {"class": "node"})
        ET.SubElement(node_attrs, "attribute", {"id": "0", "title": "type", "type": "string"})
        ET.SubElement(node_attrs, "attribute", {"id": "1", "title": "frequency", "type": "integer"})
        ET.SubElement(node_attrs, "attribute", {"id": "2", "title": "degree", "type": "integer"})

        edge_attrs = ET.SubElement(graph_el, "attributes", {"class": "edge"})
        ET.SubElement(edge_attrs, "attribute", {"id": "0", "title": "weight", "type": "float"})

        nodes_el = ET.SubElement(graph_el, "nodes")
        for node in graph.get("nodes", []):
            node_id = str(node["id"])
            label = str(node.get("label") or node_id)
            n = ET.SubElement(nodes_el, "node", {"id": node_id, "label": label})
            attvals = ET.SubElement(n, "attvalues")
            ET.SubElement(attvals, "attvalue", {
                "for": "0", "value": str(node.get("type") or "term"),
            })
            ET.SubElement(attvals, "attvalue", {
                "for": "1", "value": str(node.get("frequency") or 0),
            })
            ET.SubElement(attvals, "attvalue", {
                "for": "2", "value": str(node.get("degree") or 0),
            })

        edges_el = ET.SubElement(graph_el, "edges")
        for edge_idx, edge in enumerate(graph.get("edges", [])):
            e = ET.SubElement(
                edges_el,
                "edge",
                {
                    "id": str(edge_idx),
                    "source": str(edge["source"]),
                    "target": str(edge["target"]),
                    "weight": str(edge.get("weight") or 1),
                },
            )
            attvals = ET.SubElement(e, "attvalues")
            ET.SubElement(attvals, "attvalue", {
                "for": "0", "value": str(edge.get("weight") or 1),
            })

        return self._serialize_gexf(gexf)

    # ------------------------------------------------------------------
    # GEXF — bipartite actor-term network
    # ------------------------------------------------------------------

    def _build_bipartite_gexf(self, graph: dict[str, Any]) -> bytes:
        """Serialize a bipartite actor-term graph dict to GEXF.

        Consumes the graph dict format returned by
        :func:`~issue_observatory.analysis.network.build_bipartite_network`:

        - **Actor nodes**: ``id`` (pseudonymized author id), ``label``
          (display name), ``type`` = ``"actor"``
        - **Term nodes**: ``id`` (prefixed with ``"term:"``), ``label``
          (raw term string), ``type`` = ``"term"``
        - **Edges**: ``source`` (actor id), ``target`` (``"term:"``-prefixed
          term id), ``weight``

        The graph dict is produced by the database query in ``network.py``
        and therefore already excludes duplicate-flagged records.

        Args:
            graph: Graph dict ``{"nodes": [...], "edges": [...]}`` as returned
                by :func:`~issue_observatory.analysis.network.build_bipartite_network`.

        Returns:
            UTF-8 encoded GEXF XML bytes.
        """
        gexf, graph_el = self._make_gexf_root("Bipartite actor-term network")

        node_attrs = ET.SubElement(graph_el, "attributes", {"class": "node"})
        ET.SubElement(node_attrs, "attribute", {"id": "0", "title": "type", "type": "string"})
        ET.SubElement(node_attrs, "attribute", {"id": "1", "title": "platform", "type": "string"})

        edge_attrs = ET.SubElement(graph_el, "attributes", {"class": "edge"})
        ET.SubElement(edge_attrs, "attribute", {"id": "0", "title": "weight", "type": "float"})

        nodes_el = ET.SubElement(graph_el, "nodes")
        for node in graph.get("nodes", []):
            node_id = str(node["id"])
            label = str(node.get("label") or node_id)
            n = ET.SubElement(nodes_el, "node", {"id": node_id, "label": label})
            attvals = ET.SubElement(n, "attvalues")
            ET.SubElement(attvals, "attvalue", {
                "for": "0", "value": str(node.get("type") or "actor"),
            })
            # Actor nodes carry a platform value; term nodes get an empty string.
            ET.SubElement(attvals, "attvalue", {
                "for": "1", "value": str(node.get("platform") or ""),
            })

        edges_el = ET.SubElement(graph_el, "edges")
        for edge_idx, edge in enumerate(graph.get("edges", [])):
            e = ET.SubElement(
                edges_el,
                "edge",
                {
                    "id": str(edge_idx),
                    "source": str(edge["source"]),
                    "target": str(edge["target"]),
                    "weight": str(edge.get("weight") or 1),
                },
            )
            attvals = ET.SubElement(e, "attvalues")
            ET.SubElement(attvals, "attvalue", {
                "for": "0", "value": str(edge.get("weight") or 1),
            })

        return self._serialize_gexf(gexf)

    # ------------------------------------------------------------------
    # GEXF — public dispatch method
    # ------------------------------------------------------------------

    async def export_gexf(
        self,
        graph: dict[str, Any] | list[dict[str, Any]],
        network_type: str = "actor",
    ) -> bytes:
        """Export a pre-computed network graph dict as GEXF for Gephi.

        The *graph* argument must be a graph dict produced by one of the
        network analysis functions in
        :mod:`issue_observatory.analysis.network` — specifically
        :func:`~issue_observatory.analysis.network.get_actor_co_occurrence`,
        :func:`~issue_observatory.analysis.network.get_term_co_occurrence`, or
        :func:`~issue_observatory.analysis.network.build_bipartite_network`.
        These functions perform all database queries (including duplicate
        exclusion) before calling this method.

        Per-arena GEXF (IP2-047):
            To generate an arena-scoped GEXF file for cross-arena comparison,
            call the corresponding API endpoint with the ``arena`` query
            parameter before passing the result to this method.  For example::

                GET /analysis/{run_id}/network/actors?arena=twitter

            The ``arena`` parameter is accepted by the ``/network/actors``,
            ``/network/terms``, and ``/network/bipartite`` endpoints and is
            passed through to the underlying network analysis function, which
            applies it as a SQL filter before returning the graph dict.  This
            method only performs XML serialization; arena scoping happens at
            the database layer.

        Dispatches to one of three GEXF serializers based on ``network_type``:

        ``"actor"`` (default)
            Actor co-occurrence network.  Nodes are pseudonymised authors;
            node attributes are ``platform``, ``post_count``, and ``degree``.
            Edge weight = co-occurrence count from the database query.

        ``"term"``
            Term co-occurrence network.  Nodes are search terms; node
            attributes are ``type``, ``frequency``, and ``degree``.
            Edge weight = number of records where both terms co-occur.

        ``"bipartite"``
            Bipartite actor-term network.  Actor nodes and term nodes are
            distinguished by the ``type`` node attribute (``"actor"`` /
            ``"term"``).  Term node IDs are prefixed with ``"term:"`` to
            avoid collision with actor IDs.  Edge weight = number of records
            where that actor matched that term.

        All three output formats:

        - GEXF 1.3 namespace (``xmlns="http://gexf.net/1.3"``)
        - ``<meta>`` block with ``creator``, ``description``, and
          ``lastmodifieddate``
        - ``<attributes class="node">`` and ``<attributes class="edge">`` with
          typed attribute declarations
        - Valid, indented UTF-8 XML with an ``<?xml version="1.0" ...?>``
          declaration

        Args:
            graph: Graph dict ``{"nodes": [...], "edges": [...]}`` as returned
                by the corresponding network analysis function in
                :mod:`issue_observatory.analysis.network`.
            network_type: One of ``"actor"``, ``"term"``, or ``"bipartite"``.
                Defaults to ``"actor"``.  Raises ``ValueError`` for unknown
                values.

        Returns:
            UTF-8 encoded GEXF XML bytes.

        Raises:
            ValueError: If ``network_type`` is not one of the three accepted
                values.
        """
        # If a list of records is passed, build the graph dict from them.
        graph_dict: dict[str, Any]
        if isinstance(graph, list):
            if network_type == "actor":
                graph_dict = self._records_to_actor_graph(graph)
            elif network_type in ("term",):
                graph_dict = self._records_to_term_graph(graph)
            elif network_type in ("bipartite", "enhanced_bipartite"):
                graph_dict = self._records_to_bipartite_graph(graph)
            else:
                graph_dict = {"nodes": [], "edges": []}
        else:
            graph_dict = graph

        if network_type == "actor":
            return self._build_actor_gexf(graph_dict)
        elif network_type == "term":
            return self._build_term_gexf(graph_dict)
        elif network_type == "bipartite":
            return self._build_bipartite_gexf(graph_dict)
        elif network_type == "enhanced_bipartite":
            return self._build_bipartite_gexf(graph_dict)
        else:
            raise ValueError(
                f"Unknown network_type {network_type!r}. "
                "Choose from: 'actor', 'term', 'bipartite', 'enhanced_bipartite'."
            )

    # ------------------------------------------------------------------
    # GEXF — dynamic temporal network (GEXF 1.3 mode="dynamic")
    # ------------------------------------------------------------------

    def _build_dynamic_gexf(self, snapshots: list[dict[str, Any]]) -> bytes:
        """Serialize temporal network snapshots as a dynamic GEXF file for Gephi Timeline.

        Creates a GEXF 1.3 document with ``mode="dynamic"`` and
        ``timeformat="datetime"``.  Each node and edge is annotated with
        ``<spells>`` elements that indicate the time periods during which they
        appear.  Edge IDs encode the source-target pair plus the period index
        to ensure uniqueness across snapshots.

        Args:
            snapshots: List of snapshot dicts as returned by
                :func:`~issue_observatory.analysis.network.get_temporal_network_snapshots`.
                Each item has ``"period"`` (ISO string), ``"graph"``
                (``{"nodes": [...], "edges": [...]}``).

        Returns:
            UTF-8 encoded GEXF XML bytes with ``mode="dynamic"``.
        """
        today = datetime.now(timezone.utc).strftime("%Y-%m-%d")

        gexf = ET.Element(
            "gexf",
            {
                "xmlns": "http://gexf.net/1.3",
                "xmlns:xsi": "http://www.w3.org/2001/XMLSchema-instance",
                "xsi:schemaLocation": (
                    "http://gexf.net/1.3 http://gexf.net/1.3/gexf.xsd"
                ),
                "version": "1.3",
            },
        )
        meta = ET.SubElement(gexf, "meta", {"lastmodifieddate": today})
        ET.SubElement(meta, "creator").text = "Issue Observatory"
        ET.SubElement(meta, "description").text = "Temporal network snapshots"

        graph_el = ET.SubElement(
            gexf,
            "graph",
            {
                "mode": "dynamic",
                "defaultedgetype": "undirected",
                "timeformat": "datetime",
            },
        )

        # Collect all periods in order.
        periods: list[str] = [s["period"] for s in snapshots]

        # ------------------------------------------------------------------
        # Collect node and edge appearance data across all snapshots.
        # node_periods: node_id -> list of period ISO strings when it appears
        # edge_periods: (source, target) -> list of period ISO strings
        # ------------------------------------------------------------------
        node_periods: dict[str, list[str]] = defaultdict(list)
        node_labels: dict[str, str] = {}
        node_types: dict[str, str] = {}
        # edge key: (source, target, period) for uniqueness; (source, target) for spells
        edge_data: dict[tuple[str, str], list[tuple[str, int]]] = defaultdict(list)

        for snapshot in snapshots:
            period = snapshot["period"]
            g = snapshot.get("graph", {"nodes": [], "edges": []})
            for node in g.get("nodes", []):
                nid = str(node["id"])
                node_periods[nid].append(period)
                if nid not in node_labels:
                    node_labels[nid] = str(node.get("label") or nid)
                    node_types[nid] = str(node.get("type") or "actor")
            for edge in g.get("edges", []):
                src = str(edge["source"])
                tgt = str(edge["target"])
                weight = int(edge.get("weight") or 1)
                # Canonical edge key: sort for undirected.
                key = (min(src, tgt), max(src, tgt))
                edge_data[key].append((period, weight))

        # ------------------------------------------------------------------
        # Write nodes with <spells>.
        # ------------------------------------------------------------------
        nodes_el = ET.SubElement(graph_el, "nodes")
        for nid, appearance_periods in node_periods.items():
            n = ET.SubElement(
                nodes_el,
                "node",
                {"id": nid, "label": node_labels.get(nid, nid)},
            )
            spells_el = ET.SubElement(n, "spells")
            for p in appearance_periods:
                ET.SubElement(spells_el, "spell", {"start": p, "end": p})

        # ------------------------------------------------------------------
        # Write edges with start/end attributes (one edge element per period).
        # ------------------------------------------------------------------
        edges_el = ET.SubElement(graph_el, "edges")
        edge_counter = 0
        for (src, tgt), period_weights in edge_data.items():
            for period, weight in period_weights:
                ET.SubElement(
                    edges_el,
                    "edge",
                    {
                        "id": str(edge_counter),
                        "source": src,
                        "target": tgt,
                        "weight": str(weight),
                        "start": period,
                        "end": period,
                    },
                )
                edge_counter += 1

        return self._serialize_gexf(gexf)

    async def export_temporal_gexf(self, snapshots: list[dict[str, Any]]) -> bytes:
        """Export temporal network snapshots as dynamic GEXF for Gephi Timeline.

        Uses GEXF 1.3 ``mode="dynamic"`` with ``<spells>`` elements on nodes
        and ``start``/``end`` attributes on edges.  Each snapshot's
        ``"period"`` ISO string becomes the spell/edge start and end value,
        allowing Gephi's Timeline plugin to animate network evolution over time.

        Args:
            snapshots: List of snapshot dicts as returned by
                :func:`~issue_observatory.analysis.network.get_temporal_network_snapshots`.
                Each dict must have ``"period"`` and ``"graph"`` keys.

        Returns:
            UTF-8 encoded GEXF XML bytes.
        """
        return self._build_dynamic_gexf(snapshots)

    # ------------------------------------------------------------------
    # RIS (IP2-056)
    # ------------------------------------------------------------------

    def export_ris(self, records: list[dict[str, Any]]) -> bytes:
        """Export records in RIS format for import into Zotero, Mendeley, or EndNote.

        Each record produces one RIS entry using the following tag mapping:

        - ``TY`` — Record type (``ELEC`` — Electronic Source).
        - ``TI`` — Title (from ``title`` or first 120 chars of ``text_content``).
        - ``AB`` — Abstract / body text (``text_content``).
        - ``AU`` — Author (``author_display_name`` if available).
        - ``UR`` — URL of the source record.
        - ``PY`` — Publication year (extracted from ``published_at``).
        - ``DP`` — Database / provider (``platform:arena``).
        - ``N1`` — Note (matched search terms joined with ``; ``).
        - ``ER`` — End of record (required by the RIS spec).

        RIS format specification: Research Information Systems (RIS) file format.
        Reference managers like Zotero and Mendeley accept UTF-8 encoded ``.ris``
        files.  Empty tags are omitted to keep the output compact.

        Args:
            records: List of content record dicts (keys match ORM column names).

        Returns:
            UTF-8 encoded RIS bytes.
        """
        lines: list[str] = []

        for rec in records:
            lines.append("TY  - ELEC")

            title = rec.get("title") or ""
            if not title:
                text = rec.get("text_content") or ""
                title = text[:120].replace("\n", " ") if text else ""
            if title:
                lines.append(f"TI  - {title}")

            text_content = rec.get("text_content") or ""
            if text_content:
                # Collapse newlines to spaces for single-line RIS tag value.
                lines.append(f"AB  - {text_content[:2000].replace(chr(10), ' ')}")

            author = rec.get("author_display_name") or ""
            if author:
                lines.append(f"AU  - {author}")

            url = rec.get("url") or ""
            if url:
                lines.append(f"UR  - {url}")

            published_at = rec.get("published_at")
            if published_at is not None:
                try:
                    year = str(published_at.year) if hasattr(published_at, "year") else str(published_at)[:4]
                    lines.append(f"PY  - {year}")
                except Exception:  # noqa: BLE001
                    pass

            platform = rec.get("platform") or ""
            arena = rec.get("arena") or ""
            if platform or arena:
                dp = f"{platform}:{arena}" if arena else platform
                lines.append(f"DP  - {dp}")

            terms = rec.get("search_terms_matched") or []
            if terms:
                joined = "; ".join(str(t) for t in terms if t)
                if joined:
                    lines.append(f"N1  - Matched search terms: {joined}")

            lines.append("ER  - ")
            lines.append("")  # blank line between entries

        return "\n".join(lines).encode("utf-8")

    # ------------------------------------------------------------------
    # BibTeX (IP2-056)
    # ------------------------------------------------------------------

    def export_bibtex(self, records: list[dict[str, Any]]) -> bytes:
        """Export records as a BibTeX ``.bib`` file for LaTeX/Overleaf workflows.

        Each record produces one ``@misc`` entry.  The entry key is
        ``record_{sha256[:8]}`` derived from the ``content_hash`` field (or
        a UUID-based fallback when no hash is available).

        Field mapping:

        - ``title`` — Record title (or first 120 chars of body text).
        - ``author`` — ``author_display_name``.
        - ``howpublished`` — URL of the source record (``\\url{...}``).
        - ``year`` — Publication year extracted from ``published_at``.
        - ``note`` — Matched search terms joined with ``; ``.
        - ``annote`` — ``platform:arena`` provenance annotation.

        Special characters in BibTeX field values are minimally escaped
        (braces, backslash, ``&``, ``$``, ``%``, ``#``, ``_``) so that
        LaTeX does not reject the import.  Unicode is preserved; callers
        should use a UTF-8-aware BibTeX engine (Biber/biblatex).

        Args:
            records: List of content record dicts (keys match ORM column names).

        Returns:
            UTF-8 encoded BibTeX bytes.
        """
        import hashlib  # stdlib — always available  # noqa: PLC0415

        _LATEX_ESCAPE: list[tuple[str, str]] = [
            ("\\", "\\textbackslash{}"),
            ("{", "\\{"),
            ("}", "\\}"),
            ("$", "\\$"),
            ("&", "\\&"),
            ("%", "\\%"),
            ("#", "\\#"),
            ("_", "\\_"),
        ]

        def _tex(value: str) -> str:
            """Minimally escape *value* for safe embedding in a BibTeX field."""
            # Apply escapes in order — backslash first to avoid double-escaping.
            result = value
            for char, replacement in _LATEX_ESCAPE:
                result = result.replace(char, replacement)
            return result

        lines: list[str] = []

        for rec in records:
            content_hash = rec.get("content_hash") or ""
            if content_hash and len(content_hash) >= 8:
                entry_key = f"record_{content_hash[:8]}"
            else:
                # Fallback: SHA-256 of the URL or a UUID string.
                fallback = rec.get("url") or rec.get("id") or ""
                entry_key = "record_" + hashlib.sha256(
                    str(fallback).encode("utf-8")
                ).hexdigest()[:8]

            lines.append(f"@misc{{{entry_key},")

            title = rec.get("title") or ""
            if not title:
                text = rec.get("text_content") or ""
                title = text[:120].replace("\n", " ") if text else "Untitled"
            lines.append(f"  title = {{{_tex(title)}}},")

            author = rec.get("author_display_name") or ""
            if author:
                lines.append(f"  author = {{{_tex(author)}}},")

            url = rec.get("url") or ""
            if url:
                lines.append(f"  howpublished = {{\\url{{{url}}}}},")

            published_at = rec.get("published_at")
            if published_at is not None:
                try:
                    year = str(published_at.year) if hasattr(published_at, "year") else str(published_at)[:4]
                    lines.append(f"  year = {{{year}}},")
                except Exception:  # noqa: BLE001
                    pass

            terms = rec.get("search_terms_matched") or []
            if terms:
                joined = "; ".join(str(t) for t in terms if t)
                if joined:
                    lines.append(f"  note = {{{_tex(joined)}}},")

            platform = rec.get("platform") or ""
            arena = rec.get("arena") or ""
            if platform or arena:
                annote = f"{platform}:{arena}" if arena else platform
                lines.append(f"  annote = {{{_tex(annote)}}},")

            lines.append("}")
            lines.append("")  # blank line between entries

        return "\n".join(lines).encode("utf-8")
